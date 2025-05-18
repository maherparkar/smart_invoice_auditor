from calendar import day_abbr
from fileinput import filename
import json
from msilib.schema import Directory
import time
from cv2 import compare
import requests
import os
import re
from pdf2image import convert_from_path
from PIL import Image
import csv
import pandas as pd 
import time
import urllib.error
import urllib.parse
import requests
from flask import current_app as app
import pytesseract
from PIL import Image
import cv2
import numpy as np 
import glob
import os
import re
import json
import re
import copy
from PIL import Image, ImageFilter
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import requests
import urllib.parse
import openpyxl
from requests_html import HTMLSession

def icr_run(image_path):
    # print("sagen")
    subscription_key = ""
  
    endpoint = ""
    
    analyze_url = endpoint + "vision/v2.1/read/core/asyncBatchAnalyze?language=en"
    image_data = open(image_path, "rb").read()
    headers = {'Ocp-Apim-Subscription-Key': subscription_key,
               'Content-Type': 'application/octet-stream'}
    params = {'visualFeatures': 'Categories,Description,Color'}
    response = requests.post(
        analyze_url, headers=headers, params=params, data=image_data)
    response.raise_for_status()
    operation_url = response.headers["Operation-Location"]

    analysis = {}
    poll = True
    while (poll):
        response_final = requests.get(
            response.headers["Operation-Location"], headers=headers, params=params)
        analysis = response_final.json()
        time.sleep(1)

        if ("recognitionResults" in analysis):
            poll = False
        if ("status" in analysis and analysis['status'] == 'Failed'):
            poll = False
    text_json = analysis['recognitionResults'][0].get("lines")
    return text_json

def get_icr_data_from_image(filepath: str):
    subscription_key = ""
    ICR_ENDPOINT = ""
    ICR_URL = ICR_ENDPOINT + "vision/v2.1/read/core/asyncBatchAnalyze?language=en"
    text_recognition_url = ICR_URL
    image_data = open(filepath, "rb").read()
    headers = {'Ocp-Apim-Subscription-Key': subscription_key,
               'Content-Type': 'application/octet-stream'}
    params = {'visualFeatures': 'Categories,Description,Color', 'language': 'en',
              'detectOrientation ': 'true'}
    response = requests.post(
        text_recognition_url, headers=headers, params=params, data=image_data)
    response.raise_for_status()
    analysis = {}
    poll = True
    while poll:
        response_final = requests.get(
            response.headers["Operation-Location"], headers=headers, params=params)
        analysis = response_final.json()
        time.sleep(1)
        if "recognitionResults" in analysis:
            poll = False
        if "status" in analysis and analysis['status'] == 'Failed':
            poll = False
    text_json = analysis['recognitionResults'][0].get("lines")
    json_data=text_json
    return text_json

def change_resolution(input_image_path, output_image_path, basewidth, baseheight):
    image = cv2.imread(input_image_path, 1)
    resized_image = cv2.resize(image, (basewidth, baseheight))
    cv2.imwrite(output_image_path, resized_image)

def pdf_to_img(pdf_name, folder_name,download_folder,index=1):
    if os.path.exists(os.path.join(download_folder, str(folder_name), str(pdf_name))):
        print(46)
    i = index
    pages = convert_from_path(os.path.join(download_folder, str(folder_name), str(pdf_name)),
                              dpi=300,output_folder=r'')
    for page in pages:
        page.save(os.path.join(download_folder, str(folder_name),
                            str(pdf_name) + str(i) + ".jpg"), 'JPEG')
        change_resolution(os.path.join(download_folder, str(folder_name),
                            str(pdf_name) + str(i) + ".jpg"),  os.path.join(download_folder, str(folder_name),
                            str(pdf_name) + str(i) + ".jpg"), 2200, 2700)
        
        i += 1
    return i


def Compare_Data(lr,invoice,doc_type,filename):
    colour_dict = {}
    all_data_dict = {}

    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")

    INdia_states = ["Andaman" ,"Nicobar","Andhra","Arunachal","Assam","Bihar","Chandigarh","Chhattisgarh","Dadra" ,"Nagar Haveli","Daman","Diu",
	                "Delhi","Goa","Gujarat","Haryana","Himachal","Jammu" , "Kashmir","Jharkhand","Karnataka","Kerala","Lakshadweep","Madhya","Maharashtra","Manipur",	
                    "Meghalaya","Mizoram","Nagaland","Orissa","Puducherry" ,"Punjab","Rajasthan","Sikkim","Tamil", "Nadu","Telangana","Tripura","Uttar","Uttarakhand","West Bengal"
                    ]
    state_shorts = ["AP","AR","AS","BR","CG","GA","GJ","HR","HP","JK","JH","KA",'KL',"MP","MH","MN","ML","MZ","NL","OR","PB","RJ","SK","TN","TS","TR","UK","UP","WB","TN","TR","AN","CH","DH","DD","DL","LD","PY"]

    capital_citys = ["Agartala","Aizawl","Bengaluru","Bhopal","Bhubaneswar","Chandigarh","Chennai","Dehradun","Dispur",
                        "Gandhinagar","Gangtok","Hyderabad","Imphal","Itanagar","Jaipur","Kohima","Kolkata","Lucknow",
                        "Mumbai","Panaji","Patna"	,"Raipur","Ranchi","Shillong","Shimla","Srinagar","Jammu","Thiruvananthapuram"
                    ]
    pin = ""
    if doc_type == "label" :
        # print(lr["COD Collect amount"],"<--------->",invoice["TOTAL PRICE"])
        if "TOTAL PRICE" in invoice.keys():
            if "COD Collect amount" in lr.keys():
                if lr["COD Collect amount"].replace("Rs.","").strip(" .") == invoice["TOTAL PRICE"].replace("Rs.","").strip(" ."):
                    print("hit1",lr["COD Collect amount"].replace("Rs.","").strip(" ."),"--cargo declaration value")
                    colour_dict["cargo declaration value"] = ["GREEN",lr["COD Collect amount"].replace("Rs.","").strip(" .")]
                    all_data_dict["cargo declaration value"] = lr["COD Collect amount"].replace("Rs.","").strip(" .")
                else:
                    colour_dict["cargo declaration value"] = ["red",invoice["TOTAL PRICE"].replace("Rs.","").strip(" .")]
                    all_data_dict["cargo declaration value"] = invoice["TOTAL PRICE"].replace("Rs.","").strip(" .")
            else:
                colour_dict["cargo declaration value"] = ["red",invoice["TOTAL PRICE"].replace("Rs.","").strip(" .")]
                all_data_dict["cargo declaration value"] = invoice["TOTAL PRICE"].replace("Rs.","").strip(" .")



        if "TOTAL QTY" in invoice.keys():
            if "Quantity" in lr.keys():
                if lr["Quantity"] == invoice["TOTAL QTY"]:
                    print("hit2",lr["Quantity"])
                    colour_dict["Quantity"] = ["GREEN",lr["Quantity"]]
                    all_data_dict["Quantity"] = lr["Quantity"]
                else:
                    print(invoice["TOTAL QTY"],"hit2 ---quantity")
                    colour_dict["Quantity"] = ["red",invoice["TOTAL QTY"]]
                    all_data_dict["Quantity"] = invoice["TOTAL QTY"]
            else:
                print(invoice["TOTAL QTY"],"hit2 ---quantity")
                colour_dict["Quantity"] = ["red",invoice["TOTAL QTY"]]
                all_data_dict["Quantity"] = invoice["TOTAL QTY"]


        # print(lr["DELIVERY ADDRESS"],"<--------->",invoice["Shipping ADDRESS"])
        if "DELIVERY ADDRESS" in lr.keys() and "Shipping ADDRESS" in invoice.keys():
            address_del = lr["DELIVERY ADDRESS"].replace(" ","")
            pin_1 = ""
            if len(re.findall('-\d{6}',address_del)) == 1:
                no_pin = re.sub('-\d{6}'," ",address_del)
                # print("yes",no_pin)
                for x in lr["DELIVERY ADDRESS"].split():
                    # print(x,"===")
                    if x not in no_pin:
                        pin_1 = pin_1 + x 
                        pin_1= pin_1.strip(" ,.-")
                        pin= pin_1.strip(" ,.-")
                # print(pin_1,"pinnn")
            address_ship = invoice["Shipping ADDRESS"].replace(" ","")
            pin_2 = ""
            if len(re.findall('-\d{6}',address_ship)) == 1:
                no_pin = re.sub('-\d{6}'," ",address_ship)
                # print("yes",no_pin)
                for x in invoice["Shipping ADDRESS"].split():
                    if x not in no_pin:
                        pin_2 = pin_2 + x 
                        pin_2 = pin_2.strip(",.- ")
                        pin = pin_2 
                # print(pin_2,"pinnn")
            if pin_1.strip(" -") == pin_2.strip(" -"):
                # print("hit3--- pincode",pin_1)
                colour_dict["pincode "] = ["GREEN",pin_1]
                all_data_dict["pincode "] = pin_1
                pin = pin_1 
            else:
                colour_dict["pincode "] = ["RED",pin_2]
                all_data_dict["pincode "] = pin_2

        if "DELIVERY ADDRESS" in lr.keys() and "Shipping ADDRESS" in invoice.keys():
            address_del = lr["DELIVERY ADDRESS"].replace(" ","")
            address_ship = invoice["Shipping ADDRESS"].replace(" ","")
            state_1 = ""
            state_2 = " "
            for x in INdia_states:
                # print(x.lower(),address_del.lower(),"fghfhjvjvhj")
                if address_del.lower().__contains__(x.lower()):
                    state_1 = x
                    print(state_1,"state_1")
            for x in INdia_states:
                if x.lower() in address_ship.lower():
                    state_2 = x
                    print(state_2,"state_2")
            if state_1 == state_2 :
                print("hit4",state_1)
            else:
                for x in state_shorts:
                    if ("IN-"+x) in address_del:
                        state_1 = "IN-"+x
                for x in state_shorts:
                    if ("IN-"+x) in address_ship:
                        state_2 = "IN-"+x
                if state_1 == state_2 :
                    print("hit44 ---loss state",state_1)
                    colour_dict["loss state"] = ["GREEN",state_1]
                    all_data_dict["loss state"] = state_1
                else:
                    colour_dict["loss state"] = ["RED",state_2]
                    all_data_dict["loss state"] = state_2
        # if "DELIVERY ADDRESS" in lr.keys() and "Shipping ADDRESS" in invoice.keys():
        #     address_del = lr["DELIVERY ADDRESS"].replace(" ","")
        #     address_ship = invoice["Shipping ADDRESS"].replace(" ","")
        #     city_1 = "1"
        #     city_2 = "2"
        #     for x in capital_citys:
        #         # print(x.lower(),capital_citys.lower(),"fghfhjvjvhj")
        #         if address_del.lower().__contains__(x.lower()):
        #             city_1 = x
        #             # print(city_1,"city_1")
        #     for x in capital_citys:
        #         if x.lower() in address_ship.lower():
        #             city_2 = x
        #             # print(city_2,"city_2")
        #     if city_1 == city_2 :
        #         print("hit5",city_1)


            # else:
            #     address_del = lr["DELIVERY ADDRESS"].split()
            #     address_ship = invoice["Shipping ADDRESS"].split()
            #     for x in address_del






                
        if "DELIVERY ADDRESS name of person" in lr.keys() and "Shipping address Name of Person" in invoice.keys():
            print(lr["DELIVERY ADDRESS name of person"],"<--------++++jgj--------->",invoice["Shipping address Name of Person"])
            stringg_list = lr["DELIVERY ADDRESS name of person"].split()
            stringg_2_list = invoice["Shipping address Name of Person"].split()
            temp_stringg = copy.deepcopy(stringg_list)
            temp_stringg_2 = copy.deepcopy(stringg_2_list)
            for x in temp_stringg:
                if str(x) == "1":
                    temp_stringg.remove(x)
            for y in temp_stringg_2: 
                if str(y) == "1":
                    temp_stringg_2.remove(y)  
            final_string=""
            # print(stringg_list,"adsdsdsdsd",stringg_2_list)
            for x in stringg_list:
                for y in stringg_2_list: 
                    if str(x) == "1" and x in stringg_list :
                        stringg_list.remove(x)
                    if str(y) == "1" and y in stringg_2_list:
                        stringg_2_list.remove(y)  
                    
                    if x.lower() == y.lower():
                        # print(x,"LLLLLLLLLLLL",y)
                        if x in stringg_list:
                            stringg_list.remove(x) 
                        if x in stringg_2_list:
                            stringg_2_list.remove(x) 
            # print(temp_stringg,"0000",temp_stringg_2)
            print(stringg_list,"0000",stringg_2_list)

            if len(stringg_list)>len(stringg_2_list):
                for x in stringg_list:
                    if x == stringg_list[-1]:
                        break
                    y = stringg_list.index(x) + 1
                    # print(x,"gggg",stringg_list[y])
                    for z in stringg_2_list:
                        if (str(x + stringg_list[y]) ).lower() == z.lower():
                            # print(z,"zzzzzzzzzzzzz")
                            INdex = temp_stringg.index(x)
                            # print(INdex,"INdexINdex")
                            temp_stringg[INdex] = z
                            print(stringg_list,"temp_stringgtemp_stringg",temp_stringg)
                            del temp_stringg[INdex+1]
                            temp_stringg_2 =" ".join(temp_stringg_2)
                            final_string = " ".join(temp_stringg)
                            # print(final_string.lower(),temp_stringg_2.lower())
                            if final_string.lower() == temp_stringg_2.lower():
                                print("hit6consighnee",final_string)
                                colour_dict["consighnee"] = ["GREEN",final_string]
                                all_data_dict["consighnee"] = final_string

            elif len(stringg_2_list)>len(stringg_list):
                for x in stringg_2_list:
                    if x == stringg_2_list[-1]:
                        break
                    y = stringg_2_list.index(x) + 1
                    # print(x,"gggfsfsg",stringg_2_list[y])
                    for z in stringg_list:
                        if (str(x + stringg_2_list[y]) ).lower() == z.lower():
                            # print(z,"zzzfsfszzzzzzzzzz")
                            INdex = temp_stringg_2.index(x)
                            # print(INdex,"INdefsfsxINdex")
                            temp_stringg_2[INdex] = z
                            # print(stringg_2_list,"temp_stridvvxnggtemp_stringg",temp_stringg_2)
                            del temp_stringg_2[INdex+1]
                            final_string = " ".join(temp_stringg_2)
                            temp_stringg =" ".join(temp_stringg)
                            if final_string.lower() == temp_stringg.lower():
                                colour_dict["consighnee"] = ["GREEN",final_string]
                                all_data_dict["consighnee"] = final_string
                                print("hit66consighnee",final_string)

            elif len(temp_stringg) == len(temp_stringg_2):
                temp_stringg = " ".join(temp_stringg)
                temp_stringg_2 = " ".join(temp_stringg_2)
                if temp_stringg.lower() == temp_stringg_2.lower():
                    colour_dict["consighnee"] = ["GREEN",temp_stringg]
                    all_data_dict["consighnee"] = temp_stringg 
                    print("hit666---consighnee",temp_stringg)
            else:
                colour_dict["consighnee"] = ["red",temp_stringg_2]
                all_data_dict["consighnee"] = temp_stringg_2 
        if "consighnee" not in colour_dict.keys():
            if fuzz.ratio(lr["DELIVERY ADDRESS name of person"].replace(" ","").strip(" ,.-"),invoice["Shipping address Name of Person"].replace(" ","").strip(" ,.-")) > 90:
                colour_dict["consighnee"] = ["green",lr["DELIVERY ADDRESS name of person"].replace(" ","")]
                all_data_dict["consighnee"] = lr["DELIVERY ADDRESS name of person"] 
            else:
                colour_dict["consighnee"] = ["red",lr["DELIVERY ADDRESS name of person"].replace(" ","")]
                all_data_dict["consighnee"] = lr["DELIVERY ADDRESS name of person"] 

        if "Sold By" in lr.keys() and "Sold By" in invoice.keys():
            # print(lr["Sold By"],"<--------++++--------->",invoice["Sold By"])
            stringg_list = lr["Sold By"].split()
            stringg_2_list = invoice["Sold By"].split()
            temp_stringg = copy.deepcopy(stringg_list)
            temp_stringg_2 = copy.deepcopy(stringg_2_list)
            for x in temp_stringg:
                if str(x) == "1":
                    temp_stringg.remove(x)
            for y in temp_stringg_2: 
                if str(y) == "1":
                    temp_stringg_2.remove(y)  
            final_string=""
            # print(stringg
            final_string=""
            # print(stringg_list,"adsdsdsdsd",stringg_2_list)
            for x in stringg_list:
                for y in stringg_2_list:
                    if x.lower() == y.lower():
                        # print(x,"LLLLLLLLLLLL",y)
                        stringg_list.remove(x) 
                        stringg_2_list.remove(x) 
                        # print(temp_stringg,"0000",temp_stringg_2)
                        # print(stringg_list,"0000",stringg_2_list)

            if len(stringg_list)>len(stringg_2_list):
                for x in stringg_list:
                    if x == stringg_list[-1]:
                        break
                    y = stringg_list.index(x) + 1
                    # print(x,"gggg",stringg_list[y])
                    for z in stringg_2_list:
                        if (str(x + stringg_list[y]) ).lower() == z.lower():
                            # print(z,"zzzzzzzzzzzzz")
                            INdex = temp_stringg.index(x)
                            # print(INdex,"INdexINdex")
                            temp_stringg[INdex] = z
                            # print(stringg_list,"temp_stringgtemp_stringg",temp_stringg)
                            del temp_stringg[INdex+1]
                            temp_stringg_2 =" ".join(temp_stringg_2)
                            final_string = " ".join(temp_stringg)
                            # print(final_string.lower(),temp_stringg_2.lower())
                            if final_string.lower() == temp_stringg_2.lower():
                                print("hit7 consighner",final_string)
                                colour_dict["consighner"] = ["green",final_string]
                                all_data_dict["consighner"] = final_string 
            elif len(stringg_2_list)>len(stringg_list):
                for x in stringg_2_list:
                    if x == stringg_2_list[-1]:
                        break
                    y = stringg_2_list.index(x) + 1
                    # print(x,"gggfsfsg",stringg_2_list[y])
                    for z in stringg_list:
                        if (str(x + stringg_2_list[y]) ).lower() == z.lower():
                            # print(z,"zzzfsfszzzzzzzzzz")
                            INdex = temp_stringg_2.index(x)
                            # print(INdex,"INdefsfsxINdex")
                            temp_stringg_2[INdex] = z
                            # print(stringg_2_list,"temp_stridvvxnggtemp_stringg",temp_stringg_2)
                            del temp_stringg_2[INdex+1]
                            final_string = " ".join(temp_stringg_2)
                            temp_stringg =" ".join(temp_stringg)
                            if final_string.lower() == temp_stringg.lower():
                                print("hit77consighner",final_string)
                                colour_dict["consighner"] = ["green",final_string]
                                all_data_dict["consighner"] = final_string 
            elif len(temp_stringg) == len(temp_stringg_2):
                temp_stringg = " ".join(temp_stringg)
                temp_stringg_2 = " ".join(temp_stringg_2)
                if temp_stringg.lower() == temp_stringg_2.lower():
                    print("hit7consighner",temp_stringg)
                    colour_dict["consighner"] = ["green",temp_stringg]
                    all_data_dict["consighner"] = temp_stringg 
            else:
                colour_dict["consighner"] = ["red",temp_stringg_2]
                all_data_dict["consighner"] = temp_stringg_2 



        if "DELIVERY ADDRESS" in lr.keys() and "Shipping ADDRESS" in invoice.keys():       
            seller_address = lr["DELIVERY ADDRESS"].replace(" ","")
            seller_reg_address = invoice["Shipping ADDRESS"].replace(" ","")
            if fuzz.ratio(seller_address,seller_reg_address)>  96 :
                print(fuzz.ratio(seller_address,seller_reg_address),"hit8",lr["DELIVERY ADDRESS"])
                colour_dict["consignee address"] = ["green",lr["DELIVERY ADDRESS"]]
                all_data_dict["consignee address"] = lr["DELIVERY ADDRESS"] 
            else:
                colour_dict["consignee address"] = ["red",invoice["Shipping ADDRESS"]]
                all_data_dict["consignee address"] = invoice["Shipping ADDRESS"] 


        if "DELIVERY ADDRESS" in lr.keys() and "Shipping ADDRESS" in invoice.keys():
            # print(lr["DELIVERY ADDRESS"],"--------------------------", invoice["Shipping ADDRESS"])
            city_lr =  re.split(r"-\s{0,1}\d{6}",lr["DELIVERY ADDRESS"])
            city_lr = city_lr[0].split(",")[-1]
            city_lr = city_lr.strip(",. -")
            # print(city_lr,"cityyyyyyyyyyyyyyy")
            city_invoice =  re.split(r"-\s{0,1}\d{6}",invoice["Shipping ADDRESS"])
            city_invoice = city_invoice[0].split(",")[-1]
            city_invoice = city_invoice.strip(",. -")
            # print(city_invoice,"cityyyyyyyyyyyyyyy")

            if city_lr in city_invoice:
                print("hit9---loss city",city_lr)
                colour_dict["loss city"] = ["green",city_lr]
                all_data_dict["loss city"] = city_lr 

            elif city_invoice in city_lr:
                colour_dict["loss city"] = ["green",city_lr]
                all_data_dict["loss city"] = city_lr 
                print("hit9--loss city",city_invoice)

            else:
                print("not found --loss city")
                colour_dict["loss city"] = ["red",city_invoice]
                all_data_dict["loss city"] = city_invoice 


        product_name = ""

        if "Product" in lr.keys() and "Product"in invoice.keys():
            for x in lr["Product"].split():
                # print(x,"xxxxxxxx")
                if x.lower() in invoice["Product"].lower():
                    # print("yessssss",x)
                    product_name = product_name + " " +x  
                    colour_dict["Product"] = ["green",product_name]
                    all_data_dict["Product"] = product_name 
                else:
                    colour_dict["Product"] = ["red",seller_reg_address]
                    all_data_dict["Product"] = invoice["Product"]  

        print(product_name,"hit 10 product_name")

        state_11 = ""
        state22 = ""
        if "seller address" in lr.keys() and "Seller Registered Address" in invoice.keys():
            city_lr =  re.split(r"-\s{0,1}\d{6}",lr["seller address"])
            city_lr = city_lr[0].split(",")[-1]
            city_lr = city_lr.strip(",. -")
            for x in city_lr.split():
                for y in INdia_states:
                    if "new" in x.lower() or "pradesh"in x.lower():
                        continue
                    # print(x,"tytytytytytyty",y)
                    if fuzz.ratio(x.lower(),y.lower())>90 : 
                        # print("it is state",x)
                        state_11 = state_11 +" "+ x
                        city_lr =  re.split(r"-\s{0,1}\d{6}",lr["seller address"])
                        print(city_lr)
                        city_lr = city_lr[0].split(",")
                        print(city_lr,"2")
                        if len(city_lr[-2])>=2:
                            city_lr = city_lr[-2]
                        else:
                             city_lr = city_lr[-3]
                        # print(city_lr,"is this a city?")
            city_invoice =  re.split(r"-\s{0,1}\d{6}",invoice["Seller Registered Address"])
            city_invoice = city_invoice[0].split(",")[-1]
            city_invoice = city_invoice.strip(",. -")
            for x in city_invoice.split():
                for y in INdia_states:
                    if "new" in x.lower() or "pradesh"in x.lower():
                        continue
                    # print(x,"tytytythghghghghghytytyty",y)
                    if fuzz.ratio(x.lower(),y.lower())>90 : 
                        state22 = state22 +" "+ x
                        # print("it is stateeeee",x)
                        city_invoice =  re.split(r"-\s{0,1}\d{6}",lr["seller address"])
                        print(city_invoice)
                        city_invoice = city_invoice[0].split(",")
                        # print(city_invoice,"22222")
                        if len(city_invoice[-2])>=2:
                            city_invoice = city_invoice[-2]
                        else:
                             city_invoice = city_invoice[-3]
                        # print(city_invoice,"is this a cityyyyy?")
            if city_lr in city_invoice:
                print("hit11",city_lr)
                colour_dict["transit city"] = ["green",city_lr]
                all_data_dict["transit city"] = city_lr 

            elif city_invoice in city_lr:
                colour_dict["transit city"] = ["green",city_invoice]
                all_data_dict["transit city"] = city_invoice 
                print("hit111",city_invoice)
            else:
                colour_dict["transit city"] = ["red",city_invoice]
                all_data_dict["transit city"] = city_invoice 
                print("not found")

        if fuzz.ratio(state_11,state22) >90:
            print(":transit state")
            colour_dict["transit state"] = ["green",state_11]
            all_data_dict["transit state"] = state_11 
        else:
            colour_dict["transit state"] = ["red",state22]
            all_data_dict["transit state"] = state22 


        if "Courier Name"  in lr.keys() :
            print("hit12 -- transporter name " ,lr["Courier Name"] )
            colour_dict["transporter name"] = ["green",lr["Courier Name"]]
            all_data_dict["transporter name"] = lr["Courier Name"] 
        else:
            colour_dict["transporter name"] = ["red"," "]
            all_data_dict["transporter name"] = " " 
        # if "seller address" in lr.keys() and "Seller Registered Address" in invoice.keys(): 
        if "Invoice Date"  in invoice.keys() :
            colour_dict["Invoice Date"] = ["green",invoice["Invoice Date"].split(",")[0]]
            all_data_dict["Invoice Date"] = invoice["Invoice Date"].split(",")[0]

        # if lr["Quantity"] == invoice["TOTAL PRICE"]:
        #     print("hit3")
        # if lr["Quantity"] == invoice["TOTAL PRICE"]:
        #     print("hit4")
        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        for k, v in colour_dict.items():
            print(k,":", v)
        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        for k, v in all_data_dict.items():
            print(k,":", v)
        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        with open(str(filename)+"color"+".json", "w") as outfile:
            json.dump(all_data_dict, outfile)
        # print(colour_dict)
        # print(all_data_dict)


def NON_FA_LAbel(text_json,filename):
    correct_order = { "COD Collect amount":"","DELIVERY ADDRESS": "","DELIVERY ADDRESS name of person":"", "Courier Name": "","Courier AWB No": "","seller address": "","Sold By": "", "GSTIN No": "","Product": "","Quantity": "", "Tax Invoice :- Invoice No": "", "Tax Invoice :- GSTIN": "","Tax Invoice :- Invoice Date": "", "Tax Invoice :- Sold By": "",    "Tax Invoice :- Shipping ADDRESS": "","Tax Invoice :- Shipping address Name of Person": "", "Tax Invoice :- Billing Address": "","Tax Invoice :- Billing Address Name of Person": "",  "Tax Invoice :- Product": "","Tax Invoice :- HSN:": "","Tax Invoice :- TOTAL QTY": "","Tax Invoice :- TOTAL PRICE": "","Tax Invoice :- Seller Registered Address": ""}
    dataa = {

                "COD Collect amount":"",
                "DELIVERY ADDRESS": "",
                "DELIVERY ADDRESS name of person":"",
                "Courier Name": "",
                "Courier AWB No": "",
                "Sold By": "",
                "seller address": "",
                "GSTIN No": "",
                "Product": "",
                "Quantity": "",

                "Tax Invoice :- Invoice No": "",
                "Tax Invoice :- GSTIN": "",
                "Tax Invoice :- Invoice Date": "",
                "Tax Invoice :- Sold By": "",    
                "Tax Invoice :- Shipping ADDRESS": "",
                "Tax Invoice :- Shipping address Name of Person": "",
                "Tax Invoice :- Billing Address": "",
                "Tax Invoice :- Billing Address Name of Person": "",  
                "Tax Invoice :- Product": "",
                "Tax Invoice :- HSN:": "",
                "Tax Invoice :- TOTAL QTY": "",
                "Tax Invoice :- TOTAL PRICE": "",
                "Tax Invoice :- Seller Registered Address": ""

            }
    print("NON_FA_LAbel")
    pagesplit = []  
    lRRR = {}
    Tax_invoice = {}
    data_bill_string = ""
    data_bill = []
    for i in range(len(text_json)):
        if "Tax Invoice"in text_json[i].get("text") :
            pagesplit = text_json[i].get("boundingBox")
    for i in range(len(text_json)):
        if "COD Collect amount" in text_json[i].get("text") :
            cod_amount = text_json[i].get("text").split("COD Collect amount")[-1]
            cod_amount = cod_amount.strip(": ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    cod_amount=  cod_amount + " " + text_json[j].get("text")
                lRRR["COD Collect amount"] = cod_amount
        elif "PREPAID"   in text_json[i].get("text"):
            lRRR["COD Collect amount"] = "PREPAID  AMOUNT"

        if "DELIVERY ADDRESS"in text_json[i].get("text") :
            delivary_address = text_json[i].get("text").split("DELIVERY ADDRESS")[-1]
            delivary_address = delivary_address.strip(": ")
            for j in range(i+1,len(text_json)):
                if "Courier Name" in text_json[j].get("text") :
                    break
                if "SURFACE" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[0] ) < 15 :
                    delivary_address=  delivary_address + " " + text_json[j].get("text")
                lRRR["DELIVERY ADDRESS"] = delivary_address
            delivary_address = delivary_address.split(",")[0]
            lRRR["DELIVERY ADDRESS name of person"] = delivary_address

        if "Courier Name" in text_json[i].get("text") :
            courier_name = text_json[i].get("text").split("Courier Name")[-1]
            courier_name = courier_name.strip(": ")
            for j in range(i+1,len(text_json)):
                if "HBD" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] > text_json[i].get("boundingBox")[2] ) :
                    courier_name=  courier_name + " " + text_json[j].get("text")
            lRRR["Courier Name"] = courier_name

        if "Courier AWB No" in text_json[i].get("text") :
            courier_awb = text_json[i].get("text").split("Courier AWB No")[-1]
            courier_awb = courier_awb.strip(": ")
            for j in range(i+1,len(text_json)):
                if "CPD" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] > text_json[i].get("boundingBox")[2] ) :
                    courier_awb=  courier_awb + " " + text_json[j].get("text")
            lRRR["Courier AWB No"] = courier_awb

        if "Sold By" in text_json[i].get("text") and text_json[i].get("boundingBox")[7] < pagesplit[1] :
            seller_Add = text_json[i].get("text").split("Sold By")[-1]
            seller_Add = seller_Add.strip(": ")
            # print("Seleeeeee",seller_Add)
            for j in range(i-1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[6] ) < 150 and (text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[7] ) < 10 and "Sold By" not in text_json[j].get("text") :
                    # print(text_json[j].get("text"),"jkmghkghkbkb")
                    seller_Add = seller_Add + " " + text_json[j].get("text")
                lRRR["seller address"] = seller_Add
            seller_name = seller_Add.split(",")[0]
            lRRR["Sold By"] = seller_name

        if ("GSTIN No") in text_json[i].get("text") and text_json[i].get("boundingBox")[7] < pagesplit[1]:
            GSTIN_num = text_json[i].get("text").split("GSTIN No")[-1]
            GSTIN_num = GSTIN_num.strip(": ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    GSTIN_num=  GSTIN_num + " " + text_json[j].get("text")
                lRRR["GSTIN No"] = GSTIN_num

        rough_area_x = ["",""]
        rough_area_y = ["",""]
        rough_area = []
        quantityyy = ""
        qntity = 0
        landmark = []
        if "Qty" in text_json[i].get("text") or "oty"  in text_json[i].get("text"):
            for j in range(len(text_json)):
                qntity = 1
                if len(text_json[i].get("text"))>4:
                    for h in text_json[i].get("words"):
                        if h['text'] == "Qty" or h['text'] == "oty":
                            rough_area_x[0]=(h["boundingBox"][0])
                            rough_area_x[1]=(h["boundingBox"][2])      
                else:       
                    rough_area_x[0]=(text_json[i].get("boundingBox")[0])
                    rough_area_x[1]=(text_json[i].get("boundingBox")[2])
                for x in range(len(text_json)):
                    if len(re.findall("Tracking ID",text_json[x].get("text")))>=1:
                        landmark = text_json[x].get("boundingBox")

                if "Total" in text_json[j].get("text"):
                    if landmark[5] > text_json[j].get("boundingBox")[5]:
                        rough_area_y[0]=(text_json[j].get("boundingBox")[1])
                        rough_area_y[1]=(text_json[j].get("boundingBox")[7])
                if qntity==1:
                    rough_area = [rough_area_x[0],rough_area_y[0],rough_area_x[1],rough_area_y[0],rough_area_x[0],rough_area_y[1],rough_area_x[1],rough_area_y[1]]
            for k in range(len(text_json)):
                if (rough_area[0]-10  <= text_json[k].get("boundingBox")[2] <= rough_area[2]+5) and (rough_area[3]-10 <= text_json[k].get("boundingBox")[3] <= rough_area[5]+10):
                    if  text_json[k].get("text").isdigit():
                        quantityyy = text_json[k].get("text")
                        print("qtyyytdgx",quantityyy)
            if qntity==0:
                quantityyy = 0
            lRRR["Quantity"] = quantityyy
        
        if "Product"in text_json[i].get("text") and text_json[i].get("boundingBox")[7] < pagesplit[1] :
            product = text_json[i].get("text").split("Product")[-1]
            product = product.strip(": ")
            # print(product,"productproduct")
            for j in range(i+1,len(text_json)):
                if "Total" in text_json[j].get("text") :
                    break
                if "Qty" in text_json[j].get("text") or "oty"  in text_json[j].get("text"):
                    continue
                if(abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[0] ) < 15) or (abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[7] ) <= 25) and len(text_json[j].get('text')) >1:
                    product += ' ' + text_json[j].get('text')
                    # print(product,"productproductproduct")
                lRRR["Product"] = product

        # y = list(lRRR.keys())
        # for x in y:
        #     if len(x) < 32: 
        #         z = 32 - len(x)
        #         lRRR[x + " "*int(z)] = lRRR[x]
        #         # lRRR.update({k:(k[v]+" ")})

        #############################################################################
        #############################################################################
        #############################################################################
        #############################################################################
        #############################################################################
        #bottomhalf###xx

        if ("GSTIN:" or "GSTIN") in text_json[i].get("text") and text_json[i].get("boundingBox")[7] > pagesplit[1]:
            GSTIN_number = text_json[i].get("text").split("GSTIN")[-1]
            GSTIN_number = GSTIN_number.strip(": ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    GSTIN_number=  GSTIN_number + " " + text_json[j].get("text")
                Tax_invoice["Tax Invoice :- GSTIN"] = GSTIN_number
        sold_by_address = ""
        if "Sold By" in text_json[i].get("text") and text_json[i].get("boundingBox")[7] > pagesplit[1] :
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[6] ) <= 14 and (text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[7] ) <= 13 :
                    sold_by = text_json[j].get("text").strip(" .,")
                    Tax_invoice["Tax Invoice :- Sold By"] = sold_by
                if "Product" in text_json[j].get("text") :
                    break
                if "Description" in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <100:
                    sold_by_address = sold_by_address + " " + text_json[j].get("text")
                Tax_invoice["Tax Invoice :- Seller Registered Address"] = sold_by_address




        # if "Seller Registered Address"in text_json[i].get("text") :
        #     Seller_Registered_address = text_json[i].get("text").split("Seller Registered Address")[-1]
        #     Seller_Registered_address = Seller_Registered_address.strip(": ")
        #     for j in range(i,len(text_json)):
        #         if "Declaration" in text_json[j].get("text") :
        #             continue
        #         if "The goods sold"in text_json[j].get("text") :
        #             continue
        #         if abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[0] ) < 100 and text_json[i].get("text") !=  text_json[j].get("text") :
        #             Seller_Registered_address=  Seller_Registered_address + " " + text_json[j].get("text")
        #         Tax_invoice["Tax Invoice :- Seller Registered Address"] = Seller_Registered_address

        if "TOTAL PRICE" in text_json[i].get("text") :
            
            total_amount = text_json[i].get("text").split("TOTAL PRICE")[-1]
            total_amount = total_amount.strip(": ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                
                    total_amount=  total_amount + " " + text_json[j].get("text")
                Tax_invoice["Tax Invoice :- TOTAL PRICE"] = total_amount

        if "TOTAL QTY" in text_json[i].get("text") :
            total_qty = text_json[i].get("text").split("TOTAL QTY")[-1]
            total_qty = total_qty.strip(": ")
            for j in range(i+1,len(text_json)):
                if "TOTAL PRICE" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) and text_json[j].get("text") != text_json[i].get("text"):
                    total_qty=  total_qty + " " + text_json[j].get("text")
            Tax_invoice["Tax Invoice :- TOTAL QTY"] = total_qty

        productttt = ""
        if "Product" in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if "Shipping Charge" in text_json[j].get("text") :
                    break
                if (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[0] ) and (text_json[j].get("boundingBox")[1] > text_json[i].get("boundingBox")[7] ) :
                    if "HSN" in text_json[j].get("text"):
                        productttt = productttt + text_json[j].get("text").split("HSN")[0]
                    else :
                        productttt = productttt+ " "+ text_json[j].get("text")
            Tax_invoice["Tax Invoice :- Product"] = productttt
        lb_boundd =""

        shipping_address = ""   
        if "Shipping ADDRESS"in text_json[i].get("text") or "Shipping ADDRES"  in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if "Billing Address" in text_json[j].get("text") or "Billing Addres" in text_json[j].get("text"):
                    lb_boundd = text_json[j].get("boundingBox")[6]
                    # print("text_json[j].get(boundingBox)[6]",text_json[j].get("boundingBox")[6],"lb_bounddlb_boundd")
                if ("Product" or "Description") in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <100:
                    for x in text_json[j].get("words"):
                        if x.get("boundingBox")[4] < (int(lb_boundd)-5):
                            shipping_address = shipping_address + " " + x.get("text")
                        if x.get("boundingBox")[4] > (int(lb_boundd)-5):
                            data_bill_string = data_bill_string + " " + x.get("text")
                    # shipping_address = shipping_address + " " + text_json[j].get("text")
                Tax_invoice["Tax Invoice :- Shipping ADDRESS"] = shipping_address
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30 and abs(text_json[i].get("boundingBox")[7] - text_json[j].get("boundingBox")[1]) <15 :
                    Tax_invoice["Tax Invoice :- Shipping address Name of Person"] = text_json[j].get("text").strip(" ,.")
        # print(data_bill_string,"data_bill_stringdata_bill_string")
        billing_address = ""   
        if "Billing Address"in text_json[i].get("text") or "Billing Addres"  in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if ("Product" or "Description") in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30:
                    billing_address = billing_address + " " + text_json[j].get("text")
                Tax_invoice["Tax Invoice :- Billing Address"] = billing_address
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30 and abs(text_json[i].get("boundingBox")[7] - text_json[j].get("boundingBox")[1]) <15 :
                    Tax_invoice["Tax Invoice :- Billing Address Name of Person"] = text_json[j].get("text").strip(" ,.")

        if "Invoice No" in text_json[i].get("text") :
            invoice_no = text_json[i].get("text").split("Invoice No")[-1]
            invoice_no = invoice_no.strip(": ")
            for j in range(i+1,len(text_json)):
                if "GSTIN:" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    invoice_no=  invoice_no + " " + text_json[j].get("text")
            Tax_invoice["Tax Invoice :- Invoice No"] = invoice_no

        if "Invoice Date:" in text_json[i].get("text") or "nvoice Date" in text_json[i].get("text") :
            invoice_date = text_json[i].get("text").split("Invoice Date:")[-1]
            invoice_date = invoice_date.split("nvoice Date:")[-1]
            invoice_date = invoice_date.strip(": ")
            if "PAN" in invoice_date:
                invoice_date = invoice_date.split("PAN")[0]
            for j in range(i+1,len(text_json)):
                if "PAN" in text_json[j].get("text") :
                    break
                if "PAN:" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    invoice_date=  invoice_date + " " + text_json[j].get("text")
            Tax_invoice["Tax Invoice :- Invoice Date"] = invoice_date

        if "HSN:" in text_json[i].get("text") :
            hsnnn = text_json[i].get("text").split("HSN:")[-1]
            hsnnn = hsnnn.strip(": ")  
            if "|" in hsnnn:
                hsnnn = hsnnn.split("|")[0]
            Tax_invoice["Tax Invoice :- HSN:"] = hsnnn
    data3 = {}
    for k in Tax_invoice:
        data3[k.split("- ")[1]] = Tax_invoice[k]
    Compare_Data(lRRR,data3,"label",filename)

    # all_data_now = {1:lRRR,"Tax Invoice---->":data3}


    all_data = lRRR|Tax_invoice
    all_data_now={}
    for d in correct_order:
        if d in all_data:
            all_data_now[d]=all_data[d]
        else:
            all_data_now[d]=correct_order[d]
    # print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    # for k, v in all_data_now.items():
    #     print(k,":", v)
    # print(all_data)
    return(all_data_now)

def RVP_NON_FA_LAbel(text_json):
    print("Rvvppp non fbf type ")
    data = {
                "Invoice No":"",
                "GSTIN": "",
                "Invoice Date": "",
                "Sold By": "",
                "seller registred address": "",
                "Shipping ADDRESS": "",
                "Shipping address Name of Person": "",
                "Billing Address": "",
                "Billing Address Name of Person": "",
                "Product": "",
                "HSN:": "",    
                "TOTAL QTY": "",
                "TOTAL PRICE": ""
            }
    correct_format = {"Invoice No":"","GSTIN": "","Invoice Date": "","Sold By": "","seller registred address": "","Shipping ADDRESS": "","Shipping address Name of Person": "","Billing Address": "","Billing Address Name of Person": "","Product": "","HSN:": "","TOTAL QTY": "","TOTAL PRICE": ""}
    
    pagesplit = []
    lb_boundd = ""

    for i in range(len(text_json)):
        if ("GSTIN:" or "GSTIN") in text_json[i].get("text"):
            GSTIN_number = text_json[i].get("text").split("GSTIN")[-1]
            GSTIN_number = GSTIN_number.strip(": ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    GSTIN_number=  GSTIN_number + " " + text_json[j].get("text")
                data["GSTIN"] = GSTIN_number

        if "Sold By" in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[6] ) < 100 and (text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[7] ) < 10 :
                    sold_by = text_json[j].get("text")
                    data["Sold By"] = sold_by

        seller_registered_address = ""   
        if "Sold By"in text_json[i].get("text") :
            print(text_json[i].get("boundingBox"),"sgsgxgxgh")
            for j in range(i+1,len(text_json)):

                if ("Product" or "Description") in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <100:
                    print(text_json[j].get("text"),"sgsgxpkpkpgxgh")
                    seller_registered_address = seller_registered_address + " " + text_json[j].get("text")
                data["seller registred address"] = seller_registered_address

        if "TOTAL PRICE" in text_json[i].get("text") :
            total_amount = text_json[i].get("text").split("TOTAL PRICE")[-1]
            total_amount = total_amount.strip(": ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    total_amount=  total_amount + " " + text_json[j].get("text")
                data["TOTAL PRICE"] = total_amount

        if "TOTAL QTY" in text_json[i].get("text") :
            total_qty = text_json[i].get("text").split("TOTAL QTY")[-1]
            total_qty = total_qty.strip(": ")
            for j in range(i+1,len(text_json)):
                if "TOTAL PRICE" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) and text_json[j].get("text") != text_json[i].get("text"):
                    total_qty=  total_qty + " " + text_json[j].get("text")
            data["TOTAL QTY"] = total_qty
        product = ""
              
        if "Product" in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if "Shipping Charge" in text_json[j].get("text") :
                    break
                if (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[0] ) and (text_json[j].get("boundingBox")[1] > text_json[i].get("boundingBox")[7] ) :
                    if "HSN" in text_json[j].get("text"):
                        product = product + text_json[j].get("text").split("HSN")[0]
                    else :
                        product = product+ " "+ text_json[j].get("text")
            data["Product"] = product
        shipping_address = ""
        lb_boundd = ""

        if "Shipping ADDRESS"in text_json[i].get("text") or "Shipping ADDRES"  in text_json[i].get("text") :
            for j in range(i,len(text_json)):
                if "Billing Address"in text_json[j].get("text") or "Billing Addres"in text_json[j].get("text"):
                    lb_boundd = text_json[j].get("boundingBox")[6]
                    print(lb_boundd,"wsedddddddaq",text_json[j].get("text"))
            for j in range(i+1,len(text_json)):
                if ("Product" or "Description") in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30:
                    if lb_boundd != "" or lb_boundd != " ":
                        for x in text_json[j].get("words"):
                            if x.get("boundingBox")[4] < (int(lb_boundd)-5):
                                shipping_address = shipping_address + " " + x.get("text")
                data["Shipping ADDRESS"] = shipping_address
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30 and abs(text_json[i].get("boundingBox")[7] - text_json[j].get("boundingBox")[1]) <15 :
                    data["Shipping address Name of Person"] = text_json[j].get("text").strip(" ,.")

        billing_address = ""   
        if "Billing Address"in text_json[i].get("text") or "Billing Addres"  in text_json[i].get("text") :
            print("it is working here wawdaaff")
            for j in range(i+1,len(text_json)):
                if ("Product" or "Description") in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30:
                    billing_address = billing_address + " " + text_json[j].get("text")
                data["Billing Address"] = billing_address
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30 and abs(text_json[i].get("boundingBox")[7] - text_json[j].get("boundingBox")[1]) <15 :
                    data["Billing Address Name of Person"] = text_json[j].get("text").strip(" ,.")


        if "Invoice No" in text_json[i].get("text") :
            invoice_no = text_json[i].get("text").split("Invoice No")[-1]
            invoice_no = invoice_no.strip(": ")
            for j in range(i+1,len(text_json)):
                if "GSTIN:" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    invoice_no=  invoice_no + " " + text_json[j].get("text")
            data["Invoice No"] = invoice_no

        if "Invoice Date:" in text_json[i].get("text") :
            invoice_date = text_json[i].get("text").split("Invoice Date:")[-1]
            invoice_date = invoice_date.strip(": ")
            for j in range(i+1,len(text_json)):
                if "PAN" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    invoice_date=  invoice_date + " " + text_json[j].get("text")
            data["Invoice Date"] = invoice_date

        if "HSN:" in text_json[i].get("text") :
            hsnnn = text_json[i].get("text").split("HSN:")[-1]
            hsnnn = hsnnn.strip(": ")  
            if "|" in hsnnn:
                hsnnn = hsnnn.split("|")[0]
            data["HSN:"] = hsnnn 

    # print("############################################################")
    # for k, v in data.items():
    #     print(k,":", v)
    # print("############################################################")

    all_data_now={}
    for d in correct_format:
        if d in data:
            all_data_now[d]=data[d]
        else:
            all_data_now[d]=correct_format[d]
    return data

def RVP_type_2(text_json):

    print("RVP_type_2")   
    data3 = {}
    correct_format = {"Sold By":"","Seller Registered Address":"",	"GSTIN":"","billing ADDRESS": "","billing ADDRESS Name of Person": "","Shipping ADDRESS": "","Shipping address Name of Person": "", "Invoice Date": "","Invoice Number": "","product name": "","Quantity": "","HSN:": "","Grand Total": ""}
    
    data3 = {
                "Sold By":"",
                "Seller Registered Address":"",	
                "GSTIN":"",
                "billing ADDRESS": "",
                "billing ADDRESS Name of Person": "",
                "Shipping ADDRESS": "",
                "Shipping address Name of Person": "",
                "Invoice Date": "",
                "Invoice Number": "",
                "product name": "",
                "Quantity": "",
                "HSN:": "",    
                "Grand Total": ""

            }
          
    for i in range(len(text_json)):

        if "Ship-from Address"in text_json[i].get("text") :
            Seller_Registered_address = text_json[i].get("text").split("Ship-from Address")[-1]
            Seller_Registered_address = Seller_Registered_address.strip(": ")
            for j in range(i+1,len(text_json)):
                if "GSTIN" in text_json[j].get("text") :
                    break
                if "Invoice Number" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[0] ) < 15 :
                    Seller_Registered_address=  Seller_Registered_address + " " + text_json[j].get("text")
                data3["Seller Registered Address"] = Seller_Registered_address

        if "Sold By" in text_json[i].get("text")  :
            sold_by = text_json[i].get("text").split("Sold By")[-1]
            sold_by = sold_by.strip(": ")
            for j in range(i+1,len(text_json)):
                if "Ship-from Address" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[6] ) < 10 and (text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[7] ) < 10 :
                    sold_by = sold_by +" "+ text_json[j].get("text")
            data3["Sold By"] = sold_by
        shipping_address = ""

        if "Ship To"in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if ("Product" or "Title") in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30:
                    shipping_address = shipping_address + " " + text_json[j].get("text")
                data3["Shipping ADDRESS"] = shipping_address
                if abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <30 and abs(text_json[i].get("boundingBox")[7] - text_json[j].get("boundingBox")[1]) <15 :
                    data3["Shipping address Name of Person"] = text_json[j].get("text").strip(" ,.")

        if "Bill To"in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if ("Product" or "Title") in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[2] - text_json[j].get("boundingBox")[0]) <150:
                    shipping_address = shipping_address + " " + text_json[j].get("text")
                data3["billing ADDRESS"] = shipping_address
                if abs(text_json[i].get("boundingBox")[2] - text_json[j].get("boundingBox")[0]) <150 and abs(text_json[i].get("boundingBox")[7] - text_json[j].get("boundingBox")[1]) <=17 :
                    data3["billing ADDRESS Name of Person"] = text_json[j].get("text").strip(" ,.")
        if "HSN/SAC:" in text_json[i].get("text") or "/SAC" in text_json[i].get("text"):
            hsnnn = text_json[i].get("text").split("HSN/SAC:")[-1]
            hsnnn = hsnnn.split("/SAC:")[-1]
            hsnnn = hsnnn.strip(": ")  
            data3["HSN:"] = hsnnn 

        rough_area_x = ["",""]
        rough_area_y = ["",""]
        rough_area = []
        quantityyy = ""
        qntity = 0
        landmark = []


        if "Qty" in text_json[i].get("text"):
            for j in range(len(text_json)):
                qntity = 1
                if len(text_json[i].get("text"))>4:
                    for h in text_json[i].get("words"):
                        if h['text'] == "Qty":
                            rough_area_x[0]=(h["boundingBox"][0])
                            rough_area_x[1]=(h["boundingBox"][2])      
                else:       
                    rough_area_x[0]=(text_json[i].get("boundingBox")[0])
                    rough_area_x[1]=(text_json[i].get("boundingBox")[2])

                if "Total" in text_json[j].get("text") and text_json[j].get("boundingBox")[0] < rough_area_x[0] and len(text_json[j].get("text"))<=6:
                    rough_area_y[0]=(text_json[j].get("boundingBox")[1])
                    rough_area_y[1]=(text_json[j].get("boundingBox")[7])

                if qntity==1:
                    rough_area = [rough_area_x[0],rough_area_y[0],rough_area_x[1],rough_area_y[0],rough_area_x[0],rough_area_y[1],rough_area_x[1],rough_area_y[1]]

            for k in range(len(text_json)):
                if (rough_area[0]-50  <= text_json[k].get("boundingBox")[2] <= rough_area[2]+5) and (rough_area[3]-10 <= text_json[k].get("boundingBox")[3] <= rough_area[5]+10):
                    quantityyy = text_json[k].get("text")
                if quantityyy == "":
                    quantityyy =1 
            if qntity==0:
                quantityyy = 0
            data3["Quantity"] = quantityyy


        product = ""
        if "Title" in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if "Warranty" in text_json[j].get("text") :
                    break
                if "IGST:" in text_json[j].get("text") :
                    break
                if  text_json[j].get("boundingBox")[1] > 250 and abs(text_json[j].get("boundingBox")[0] - text_json[i].get("boundingBox")[0]) <15 :
                    product = product + " " + text_json[j].get("text")
            data3["product name "] = product

        if "Grand Total" in text_json[i].get("text")  :
            grand_total = text_json[i].get("text").split("Grand Total")[-1]
            grand_total = grand_total.strip(": ")
            for j in range(i+1,len(text_json)):
                if "Authorized Signatory" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[1] ) < 100 and abs(text_json[j].get("boundingBox")[0] > text_json[i].get("boundingBox")[6] )  :
                    grand_total = grand_total +" "+ text_json[j].get("text")
                    grand_total = re.sub(re.sub(r'\d{2,}.\d{0,2}',"",grand_total),"",grand_total)
            data3["Grand Total"] = grand_total

        if "GSTIN" in text_json[i].get("text"):
            GSTIN_number = text_json[i].get("text").split("GSTIN")[-1]
            GSTIN_number = GSTIN_number.strip(":- ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    GSTIN_number=  GSTIN_number + " " + text_json[j].get("text")
                data3["GSTIN"] = GSTIN_number

        if "Invoice Number" in text_json[i].get("text") :
            invoice_no = text_json[i].get("text").split("Invoice Number")[-1]
            invoice_no = invoice_no.strip(": ")
            for j in range(i+1,len(text_json)):
                if "GSTIN:" in text_json[j].get("text") :
                    break
                if abs(text_json[j].get("boundingBox")[1] - text_json[i].get("boundingBox")[3] ) < 10 and (text_json[j].get("boundingBox")[0] < text_json[i].get("boundingBox")[2] ) :
                    invoice_no=  invoice_no + " " + text_json[j].get("text")
                data3["Invoice Number"] = invoice_no
        if "Original Invoice Number"in text_json[i].get("text") :
            for j in range(i+1,len(text_json)):
                if "Reason Of Issuance" in text_json[j].get("text") :
                    break
                if abs(text_json[i].get("boundingBox")[7] - text_json[j].get("boundingBox")[1]) <=20 and abs(text_json[i].get("boundingBox")[0] - text_json[j].get("boundingBox")[0]) <=20:
                    invoice_no_original =  text_json[j].get("text")
                    data3["Invoice Number"] = invoice_no_original

        if "Invoice Date:" in text_json[i].get("text") :
            invoice_date = text_json[i].get("text").split("Invoice Date:")[-1]
            invoice_date = invoice_date.strip(": ")
            data3["Invoice Date"] = invoice_date
        if data3["Invoice Date"] == "" :
            if "Order Date" not in text_json[i].get("text"):
                if len(re.findall(r'\d{1,2}-\d{1,2}-\d{0,4}',text_json[i].get("text"))) == 1:
                        print("woillll workkkkk",text_json[i].get("text"))
                        data3["Invoice Date"] = re.sub(re.sub(r'\d{1,2}-\d{1,2}-\d{,4}',"",text_json[i].get("text")),"",text_json[i].get("text"))
                        print(data3["Invoice Date"],"invoiceeee dattteeeee")

    all_data_now={}
    for d in correct_format:
        if d in data3:
            all_data_now[d]=data3[d]
        else:
            all_data_now[d]=correct_format[d]
    return data3

def rvp_screenshot(text_json):
    
    data = {}
    for i in range(len(text_json)):
        if "Vendor Tracking ID"in text_json[i].get("text") or "Vendor"in text_json[i].get("text") :
            vendor_tracking_id = text_json[i].get("text").split("Vendor Tracking ID")[-1]
            vendor_tracking_id = vendor_tracking_id.strip(": ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[7] - text_json[i].get("boundingBox")[7] ) < 10 and abs(text_json[i].get("boundingBox")[2] - text_json[j].get("boundingBox")[0] ) <=110 :
                    vendor_tracking_id=  vendor_tracking_id + " " + text_json[j].get("text")
                data["Vendor Tracking ID"] = vendor_tracking_id
        if "Merchant Shipment ID"in text_json[i].get("text") or "Merchant Shipment ID"in text_json[i].get("text") :
            vendor_tracking_id = text_json[i].get("text").split("Merchant Shipment ID")[-1]
            vendor_tracking_id = vendor_tracking_id.strip(": ")
            for j in range(i+1,len(text_json)):
                if abs(text_json[j].get("boundingBox")[7] - text_json[i].get("boundingBox")[7] ) < 10 and abs(text_json[i].get("boundingBox")[2] - text_json[j].get("boundingBox")[0] ) <=100 :
                    vendor_tracking_id=  vendor_tracking_id + " " + text_json[j].get("text")
                data["Merchant Shipment ID"] = vendor_tracking_id
    return data





def Recognise_page(text_json,filename):
    print("entered page recognision")
    tally = ["","",""]
    temp = []
    total_num_1 = 0
    final_data = {}

    for i in range(len(text_json)):
        # print(text_json[i].get("text"))
        # print(text_json[i].get("boundingBox"))
        # print("######################################")

        if "COD Collect amount" in  text_json[i].get("text"):
            tally[0] = 1
        if "Handover to" in  text_json[i].get("text"):
            tally[1] = 1
        if len(text_json) > 1 and  text_json[i].get("boundingBox")[0] < 300 and text_json[i].get("boundingBox")[1] < 240 : 
            temp.append(text_json[i].get("text"))
        if len(temp)<1 :
            tally[2] = 1 
    for x in tally:
        if x == 1:
                total_num_1 +=1
    if total_num_1 >= 2:
        final_data = NON_FA_LAbel(text_json,filename)
        # if len(final_data)>1:
        #     with open(str(filename)+"NonFafLabel"+".json", "w") as outfile:
        #         json.dump(final_data, outfile)
    tally = ["","",""]
    if total_num_1 < 2:
        for i in range(len(text_json)):
            if "Tax Invoice" in  text_json[i].get("text") and text_json[i].get("boundingBox")[0] < 250 and text_json[i].get("boundingBox")[1] < 240  :
                tally[0] = 1
            for j in range(len(text_json)):
                if "Order Id" in  text_json[j].get("text"):
                    if "Tax Invoice" in  text_json[i].get("text"):
                        if abs(text_json[i].get("boundingBox")[1] - text_json[j].get("boundingBox")[1]) <= 25 and text_json[i].get("boundingBox")[1] < text_json[j].get("boundingBox")[0]:
                            tally[1] = 1
            if "TOTAL QTY" in  text_json[i].get("text"):
                tally[2] = 1 
        for x in tally:
            if x == 1:
                    total_num_1 +=1
        if total_num_1 >= 2:
            final_data =RVP_NON_FA_LAbel(text_json)
            print("page recognised -- Rvp non fa")
            if len(final_data)>1:
                with open(str(filename)+"RvpNonFa"+".json", "w") as outfile:
                    json.dump(final_data, outfile)
    total_num_1 = 0
    tally = ["","",""]
    temp = []
    if total_num_1 < 2:
        for i in range(len(text_json)):
            if "Grand Total" in  text_json[i].get("text"):
                tally[0] = 1
            if "Ship To" in  text_json[i].get("text"):
                tally[1] = 1
            if "Bill To" in  text_json[i].get("text") :
                tally[2] = 1 
        for x in tally:
            if x == 1:
                    total_num_1 +=1
        if total_num_1 >= 2:
            final_data =RVP_type_2(text_json)
            print("page recognised -- Rvp type 2 ")
            if len(final_data)>1:
                with open(str(filename)+"RvpType2"+".json", "w") as outfile:
                    json.dump(final_data, outfile)
    for i in range(len(text_json)):
        if "Vendor Tracking ID" in text_json[i].get("text"):
            final_data = rvp_screenshot(text_json)
            with open(str(filename)+"rvp_screenshot"+".json", "w") as outfile:
                json.dump(final_data, outfile)          
    if len(final_data) == 0 :
        print("page not recognised")
        final_data = "no dataa!!"

    return final_data

def excel_out(files:list):
    sample = []
    path = ""
    fl = ""

    for fl in files:
        if fl.__contains__('RvpNonFaf'):
            path ='RvpNonFaf.xlsx'
            fl = "rvpnonfaf_head_Data.json"
        elif fl.__contains__('RvpType2'):
            path ='RvpType2.xlsx'
            fl = "RvpType2_head_Data.json"
        elif fl.__contains__('NonFafLabel'):
            path ='NonFafLabel.xlsx'
            fl = "NonFafLabel_head_Data.json"
        elif fl.__contains__('color'):
            path ='color.xlsx'
            fl = "color.json"
        f = open(fl,)
        # print(fl,"typeeeeeeeeeeee",path)
        json_dic=json.load(f)   
        sample.append(json_dic) 
    
    # df = pd.DataFrame(sample)
    # df.to_excel(path,index=False)
        mbk = openpyxl.load_workbook(path,read_only=False)
        sht = mbk['Sheet1']
        sht.cell(row=1, column=1).value='Filename'
        c=2
    for jd in json_dic:
        sht.cell(row=1, column=c).value=jd
        c +=1
    mbk.save(path)
    p = 2
    for fl in files:
        f = open(fl,)
        json_dic=json.load(f)    
        mbk = openpyxl.load_workbook(path,read_only=False)
        sht = mbk['Sheet1']
        sht.cell(row=p, column=1).hyperlink=fl.split('.json')[0]+'.pdf'
        sht.cell(row=p, column=1).value=fl.split('.json')[0]+'.pdf'
        c=2
        for jd in json_dic:
            try:
                sht.cell(row=p, column=c).value=str(json_dic[jd])
                c +=1
            except:
                c +1
                print('unexpected value',json_dic[jd])
        p +=1
        mbk.save(path)

# def this_json(Directory):
#     for filename in os.listdir(Directory):
#         f = os.path.join(Directory, filename)
#         if os.path.isfile(f) and filename.endswith('.jpg'):
#     this_dic={}
#     json_data = get_icr_data_from_image(img)
#     # for i in range(len(json_data)):
#     #     print(json_data[i].get('text'))
#     #     print(json_data[i].get('boundingBox'))
#     #     # print("**********")
#     this_dic[img]= json_data
#     with open("all_data"+img[0].split('_'+str(k_in)+'A.jpg')[0]+".json", "w") as outfile:
#         json.dump(this_dic, outfile)

def Initiate_Process(Directory):
    for filename in os.listdir(Directory):
        if  filename.endswith('.pdf'):
            # print(filename,"xxxxxxxxxxxxxxxxxxxx")
            pdf_to_img(f'{filename}','',r'')
    for filename in os.listdir(Directory):
        if filename.endswith(".PNG"):
            # print(filename,"xxxxxxxxxxxxxxxxxxxx")
            im = Image.open(filename)
            name=f'{filename}'+'.jpg'
            rgb_im = im.convert('RGB')
            rgb_im.save(name)
            # print(os.path.join(Directory, filename))
            continue
        else:
            continue

    for filename in os.listdir(Directory):
        this_dic={}
        f = os.path.join(Directory, filename)
        if os.path.isfile(f) and filename.endswith('.jpg'):
            print(filename,"xxxxxxxxxxxxxxxxxxxx")
            text_json = icr_run(f) 
            # this_dic[f]= text_json
            # with open("all_data_"+ filename +".json", "w") as outfile:
            #     json.dump(this_dic, outfile)
            final_data = Recognise_page(text_json,filename)
            if type(final_data) == dict:
                y = list(final_data.keys())
                for x in y:
                    # print(x,"Xxxxxxx")
                    if len(str(x)) < 47: 
                        z = 47 - len(str(x))
                        # print("zzzzz",z)
                        final_data[(str(x)) + " "*(int(z))] = final_data[x]
                        del final_data[x]
                print("##############################################")            
                for k, v in final_data.items():
                    print(k,":", v)
                print("##############################################")

Directory = ""
files=glob.glob('*.json')

# files.remove('NonFafLabel_head_Data.json')
# files.remove('rvpnonfaf_head_Data.json')
# files.remove('RvpType2_head_Data.json')

Initiate_Process(Directory)
# excel_out(files) 
